from openpyxl.styles import Border, Side, Font
from openpyxl.utils import get_column_letter

def ajustar_colunas(aba):
    for coluna in aba.columns:
            max_length = 0
            coluna = [cell for cell in coluna]
            for cell in coluna:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max(max_length, len(str(cell.value))) + 2
            aba.column_dimensions[get_column_letter(coluna[0].column)].width = adjusted_width
def ajustar_bordas(planilha):

    for sheet_name in planilha.sheetnames:
        worksheet = planilha[sheet_name]
        
        for col_num in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))